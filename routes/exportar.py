# routes/exportar.py

import io
import openpyxl

from flask import Blueprint, request, send_file, jsonify
from flask_login import login_required, current_user
from openpyxl.styles import Font, Alignment, PatternFill
from utils import supabase

exportar_bp = Blueprint('exportar_bp', __name__)


# ==============================
# HELPERS
# ==============================

def estilizar_cabecalho_excel(ws, cor="007BFF"):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def ajustar_texto_planilha(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def gerar_resposta_excel(wb, nome_arquivo):
    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)

    return send_file(
        file_buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def verificar_dono_turma(turma_id):
    res, _ = supabase.table("turmas") \
        .select("id") \
        .eq("id", str(turma_id)) \
        .eq("user_id", current_user.id) \
        .execute()

    return len(res[1]) > 0


def buscar_nome_turma(turma_id):
    res, _ = supabase.table("turmas") \
        .select("nome") \
        .eq("id", str(turma_id)) \
        .eq("user_id", current_user.id) \
        .single() \
        .execute()

    return res[1]["nome"] if res[1] else "Turma"


def buscar_alunos_da_turma(turma_id):
    res, _ = supabase.table("turmas_alunos") \
        .select("alunos(id, nome_completo, matricula)") \
        .eq("turma_id", str(turma_id)) \
        .execute()

    alunos = []

    for item in res[1] or []:
        aluno = item.get("alunos")
        if aluno:
            alunos.append(aluno)

    alunos.sort(key=lambda a: a.get("nome_completo") or "")
    return alunos


def nome_periodo(periodo_tipo, periodo_numero):
    mapa = {
        "avaliacao": "Avaliação",
        "bimestre": "Bimestre",
        "trimestre": "Trimestre",
        "semestre": "Semestre"
    }

    return f"{periodo_numero}ª {mapa.get(periodo_tipo, 'Avaliação')}"


# ==============================
# FREQUÊNCIA DO DIA
# ==============================

@exportar_bp.route("/exportar/turma/<uuid:turma_id>/frequencia", methods=["GET"])
@login_required
def exportar_frequencia_dia_excel(turma_id):
    try:
        data = request.args.get("data")

        if not data:
            return jsonify({"error": "Informe a data da chamada."}), 400

        if not verificar_dono_turma(turma_id):
            return jsonify({"error": "Acesso negado."}), 403

        turma_nome = buscar_nome_turma(turma_id)

        dados, _ = supabase.rpc(
            "get_frequencia_para_exportar",
            {
                "p_turma_id": str(turma_id),
                "p_data": data
            }
        ).execute()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Chamada {data}"

        ws.append(["Aluno", "Status"])

        for row in dados[1]:
            status = "Presente" if row["presente"] is True else "Falta"
            ws.append([row["nome_completo"], status])

        estilizar_cabecalho_excel(ws, "007BFF")

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18

        ajustar_texto_planilha(ws)

        return gerar_resposta_excel(
            wb,
            f"Chamada_{turma_nome}_{data}.xlsx"
        )

    except Exception as e:
        return jsonify({
            "error": "Erro ao gerar frequência do dia.",
            "details": str(e)
        }), 500


# ==============================
# FREQUÊNCIA GERAL COM FILTRO
# ==============================

@exportar_bp.route("/exportar/turma/<uuid:turma_id>/geral", methods=["GET"])
@login_required
def exportar_frequencia_geral_excel(turma_id):
    try:
        if not verificar_dono_turma(turma_id):
            return jsonify({"error": "Acesso negado."}), 403

        turma_nome = buscar_nome_turma(turma_id)
        data_inicio = request.args.get("inicio")
        data_fim = request.args.get("fim")

        alunos = buscar_alunos_da_turma(turma_id)

        freq_query = supabase.table("frequencia") \
            .select("aluno_id, presente, data") \
            .eq("turma_id", str(turma_id))

        if data_inicio:
            freq_query = freq_query.gte("data", data_inicio)

        if data_fim:
            freq_query = freq_query.lte("data", data_fim)

        freq_res, _ = freq_query.order("data", desc=False).execute()
        frequencias = freq_res[1] or []

        datas_ordenadas = sorted(list({f["data"] for f in frequencias}))

        mapa_freq = {}

        for item in frequencias:
            aluno_id = item.get("aluno_id")

            if aluno_id not in mapa_freq:
                mapa_freq[aluno_id] = {}

            mapa_freq[aluno_id][item["data"]] = item.get("presente")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Frequência Geral"

        ws.append(["Aluno"] + datas_ordenadas + ["Presenças", "Faltas", "Total", "% Presença"])

        for aluno in alunos:
            aluno_id = aluno["id"]
            nome = aluno.get("nome_completo") or ""

            row = [nome]
            total_presencas = 0
            total_faltas = 0

            for data in datas_ordenadas:
                status = mapa_freq.get(aluno_id, {}).get(data)

                if status is True:
                    row.append("P")
                    total_presencas += 1
                elif status is False:
                    row.append("F")
                    total_faltas += 1
                else:
                    row.append("-")

            total = total_presencas + total_faltas
            porcentagem = round((total_presencas / total) * 100, 1) if total > 0 else 0

            row.extend([
                total_presencas,
                total_faltas,
                total,
                f"{porcentagem}%"
            ])

            ws.append(row)

        estilizar_cabecalho_excel(ws, "007BFF")

        ws.column_dimensions["A"].width = 35

        for col in range(2, len(datas_ordenadas) + 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

        ajustar_texto_planilha(ws)

        nome_arquivo = f"Frequencia_Geral_{turma_nome}.xlsx"

        if data_inicio and data_fim:
            nome_arquivo = f"Frequencia_{turma_nome}_{data_inicio}_a_{data_fim}.xlsx"

        return gerar_resposta_excel(wb, nome_arquivo)

    except Exception as e:
        return jsonify({
            "error": "Erro ao gerar relatório geral.",
            "details": str(e)
        }), 500


# ==============================
# NOTAS MODULARES
# ==============================

def buscar_dados_notas_modulares(turma_id):
    alunos = buscar_alunos_da_turma(turma_id)

    av_res, _ = supabase.table("avaliacoes") \
        .select("id, nome, data, nota_maxima, periodo_tipo, periodo_numero, categoria, peso") \
        .eq("turma_id", str(turma_id)) \
        .eq("user_id", current_user.id) \
        .order("periodo_numero", desc=False) \
        .order("created_at", desc=False) \
        .execute()

    avaliacoes = av_res[1] or []

    if not avaliacoes:
        return alunos, [], {}

    av_ids = [av["id"] for av in avaliacoes]

    notas_res, _ = supabase.table("notas") \
        .select("aluno_id, avaliacao_id, valor") \
        .in_("avaliacao_id", av_ids) \
        .eq("user_id", current_user.id) \
        .execute()

    notas = notas_res[1] or []

    notas_map = {}

    for nota in notas:
        aluno_id = nota.get("aluno_id")
        avaliacao_id = nota.get("avaliacao_id")

        if aluno_id not in notas_map:
            notas_map[aluno_id] = {}

        notas_map[aluno_id][avaliacao_id] = nota.get("valor")

    grupos = {}

    for av in avaliacoes:
        periodo_tipo = av.get("periodo_tipo") or "avaliacao"
        periodo_numero = av.get("periodo_numero") or 1
        chave = f"{periodo_tipo}_{periodo_numero}"

        if chave not in grupos:
            grupos[chave] = {
                "chave": chave,
                "periodo_tipo": periodo_tipo,
                "periodo_numero": periodo_numero,
                "nome_periodo": nome_periodo(periodo_tipo, periodo_numero),
                "avaliacoes": []
            }

        grupos[chave]["avaliacoes"].append(av)

    periodos = list(grupos.values())
    periodos.sort(key=lambda p: (p["periodo_tipo"], p["periodo_numero"]))

    return alunos, periodos, notas_map


def calcular_media_ponderada(aluno_id, avaliacoes, notas_map):
    soma_pesos = 0
    soma_ponderada = 0
    total_lancadas = 0

    for av in avaliacoes:
        nota = notas_map.get(aluno_id, {}).get(av["id"])

        if nota is None:
            continue

        nota_maxima = float(av.get("nota_maxima") or 10)
        peso = float(av.get("peso") or 1)

        if nota_maxima <= 0:
            continue

        nota_normalizada = (float(nota) / nota_maxima) * 10

        soma_ponderada += nota_normalizada * peso
        soma_pesos += peso
        total_lancadas += 1

    if soma_pesos == 0:
        return None, total_lancadas

    return round(soma_ponderada / soma_pesos, 2), total_lancadas


@exportar_bp.route("/exportar/turma/<uuid:turma_id>/notas", methods=["GET"])
@login_required
def exportar_notas_excel(turma_id):
    try:
        if not verificar_dono_turma(turma_id):
            return jsonify({"error": "Acesso negado."}), 403

        turma_nome = buscar_nome_turma(turma_id)

        alunos, periodos, notas_map = buscar_dados_notas_modulares(turma_id)

        wb = openpyxl.Workbook()
        ws_resumo = wb.active
        ws_resumo.title = "Resumo de Notas"

        ws_resumo.append(["Aluno"] + [p["nome_periodo"] for p in periodos] + ["Média Geral"])

        for aluno in alunos:
            aluno_id = aluno["id"]
            row = [aluno.get("nome_completo") or ""]

            medias_periodo = []

            for periodo in periodos:
                media, _ = calcular_media_ponderada(
                    aluno_id,
                    periodo["avaliacoes"],
                    notas_map
                )

                row.append(media if media is not None else "-")

                if media is not None:
                    medias_periodo.append(media)

            media_geral = round(sum(medias_periodo) / len(medias_periodo), 2) if medias_periodo else "-"
            row.append(media_geral)

            ws_resumo.append(row)

        estilizar_cabecalho_excel(ws_resumo, "6610f2")
        ws_resumo.column_dimensions["A"].width = 35

        for col in range(2, len(periodos) + 3):
            ws_resumo.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

        ajustar_texto_planilha(ws_resumo)

        for periodo in periodos:
            nome_aba = periodo["nome_periodo"][:31]
            ws = wb.create_sheet(title=nome_aba)

            avaliacoes = periodo["avaliacoes"]

            cabecalho = ["Aluno"]

            for av in avaliacoes:
                cabecalho.append(f"{av['nome']} / {av.get('nota_maxima') or 10}")

            cabecalho.extend(["Notas Lançadas", "Média Ponderada"])

            ws.append(cabecalho)

            for aluno in alunos:
                aluno_id = aluno["id"]
                row = [aluno.get("nome_completo") or ""]

                for av in avaliacoes:
                    nota = notas_map.get(aluno_id, {}).get(av["id"])
                    row.append(nota if nota is not None else "-")

                media, total_lancadas = calcular_media_ponderada(
                    aluno_id,
                    avaliacoes,
                    notas_map
                )

                row.append(f"{total_lancadas}/{len(avaliacoes)}")
                row.append(media if media is not None else "-")

                ws.append(row)

            ws.append([])
            ws.append(["Detalhes das Avaliações"])
            ws.append(["Nome", "Categoria", "Data", "Nota Máxima", "Peso"])

            for av in avaliacoes:
                ws.append([
                    av.get("nome") or "",
                    av.get("categoria") or "",
                    av.get("data") or "",
                    av.get("nota_maxima") or "",
                    av.get("peso") or ""
                ])

            estilizar_cabecalho_excel(ws, "6610f2")

            ws.column_dimensions["A"].width = 35

            for col in range(2, len(cabecalho) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

            ajustar_texto_planilha(ws)

        return gerar_resposta_excel(
            wb,
            f"Notas_Modulares_{turma_nome}.xlsx"
        )

    except Exception as e:
        return jsonify({
            "error": "Erro ao gerar boletim de notas.",
            "details": str(e)
        }), 500


# ==============================
# DIÁRIO DE BORDO
# ==============================

@exportar_bp.route("/exportar/diario", methods=["GET"])
@login_required
def exportar_diario_excel():
    try:
        turma_id = request.args.get("turma_id")
        aluno_id = request.args.get("aluno_id")
        data_inicio = request.args.get("data_inicio")
        data_fim = request.args.get("data_fim")

        query = supabase.table("diario") \
            .select(
                "titulo, conteudo, data_referencia, created_at, "
                "turmas(nome), alunos(nome_completo)"
            ) \
            .eq("user_id", current_user.id)

        if turma_id:
            query = query.eq("turma_id", turma_id)

        if aluno_id:
            query = query.eq("aluno_id", aluno_id)

        if data_inicio:
            query = query.gte("data_referencia", data_inicio)

        if data_fim:
            query = query.lte("data_referencia", data_fim)

        res, _ = query.order("data_referencia", desc=False).execute()
        diarios = res[1] or []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diário de Bordo"

        ws.append([
            "Data da Ocorrência",
            "Título",
            "Turma",
            "Aluno",
            "Conteúdo",
            "Registrado em"
        ])

        for item in diarios:
            turma = item.get("turmas") or {}
            aluno = item.get("alunos") or {}

            ws.append([
                item.get("data_referencia") or "",
                item.get("titulo") or "",
                turma.get("nome") or "",
                aluno.get("nome_completo") or "",
                item.get("conteudo") or "",
                item.get("created_at") or ""
            ])

        estilizar_cabecalho_excel(ws, "17a2b8")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 28

        ajustar_texto_planilha(ws)

        nome_arquivo = "Diario_de_Bordo.xlsx"

        if data_inicio and data_fim:
            nome_arquivo = f"Diario_de_Bordo_{data_inicio}_a_{data_fim}.xlsx"
        elif data_inicio:
            nome_arquivo = f"Diario_de_Bordo_desde_{data_inicio}.xlsx"
        elif data_fim:
            nome_arquivo = f"Diario_de_Bordo_ate_{data_fim}.xlsx"

        return gerar_resposta_excel(wb, nome_arquivo)

    except Exception as e:
        return jsonify({
            "error": "Erro ao gerar relatório do diário.",
            "details": str(e)
        }), 500